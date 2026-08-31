from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config.config import OFFER_URL
from utils.utils import console_ok


def parse_brl_to_float(var_str: str) -> float:
    if not var_str or not isinstance(var_str, str):
        return 0.0
    clean_var = var_str.replace(".", "").replace(",", ".").strip()
    try:
        return float(clean_var)
    except ValueError:
        return 0.0


def export_csv_offers(data: dict):
    console_ok("Exporting offers...")
    wb = Workbook()

    sheet = wb.active
    headers = ["ID", "NOME EMISSOR", "CNPJ EMISSOR", "NOME COORDENADOR LIDER", "CNPJ COORDENADOR LIDER",
               "TIPO REQUERIMENTO", "TIPO DE OFERTA", "STATUS", "NUMERO PROCESSO", "NUMERO REGISTROS", "VALOR EM REAIS",
               "POSSUI BOOK", "DATA", "LINK"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    sheet.row_dimensions[1].height = 28
    link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
    registros = data.get("registros", [])

    valores_numericos = []
    ofertas_com_book = 0

    for offer in registros:
        offer_id = offer.get("idRequerimento")
        raw_valor = offer.get("valorTotalEmReais")
        valor_float = parse_brl_to_float(raw_valor)
        if valor_float > 0:
            valores_numericos.append(valor_float)
        if offer.get("possuiBook") in [True, "Sim", "S", "true"]:
            ofertas_com_book += 1

        url = f"{OFFER_URL}{offer_id}"
        row = [offer.get("idRequerimento"), offer.get("nomeEmissor"), offer.get("cnpjEmissor"),
               offer.get("nomeCoordenadorLider"), offer.get("cnpjCoordenadorLider"), offer.get("nomeTipoRequerimento"),
               offer.get("nomeValorMobiliario"), offer.get("statusDaOferta"), offer.get("numeroProcesso"),
               offer.get("numeroRegistro"), offer.get("valorTotalEmReais"), offer.get("possuiBook"), offer.get("data"),
               "Link Oferta" if offer_id else "N/A"]
        sheet.append(row)

        if offer_id:
            current_row = sheet.max_row
            link_cell = sheet[f"N{current_row}"]
            link_cell.hyperlink = url
            link_cell.font = link_font
            link_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_ofertas = len(registros)
    soma_valores = sum(valores_numericos)
    max_valor = max(valores_numericos) if valores_numericos else 0.0
    min_valor = min(valores_numericos) if valores_numericos else 0.0
    valor_medio = (soma_valores / len(valores_numericos)) if valores_numericos else 0.0

    sheet.append([])
    sheet.append([])

    summary_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_font = Font(name="Calibri", size=11, bold=True)
    summary_rows = [["RESUMO ESTATÍSTICO", ""], ["Total de Ofertas:", total_ofertas],
        ["Volume Total (R$):", f"{soma_valores:,.2f}"], ["Ticket Médio (R$):", f"{valor_medio:,.2f}"],
        ["Maior Oferta (R$):", f"{max_valor:,.2f}"], ["Menor Oferta (R$):", f"{min_valor:,.2f}"],
        ["Ofertas com Bookbuilding:",
         f"{ofertas_com_book} ({(ofertas_com_book / total_ofertas * 100):.1f}%)" if total_ofertas else "0"]]
    for title, val in summary_rows:
        sheet.append([title, val])
        r_idx = sheet.max_row
        sheet[f"A{r_idx}"].font = summary_font
        if title == "RESUMO ESTATÍSTICO":
            sheet[f"A{r_idx}"].fill = summary_header_fill
            sheet[f"B{r_idx}"].fill = summary_header_fill

    for col in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
    sheet.freeze_panes = "A2"
    wb.save("offers.xlsx")
    print("Saved offers.xlsx")
