from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def export_csv_offers(data: dict):
    wb = Workbook()

    sheet = wb.active
    sheet["A1"] = "ID"
    sheet["B1"] = "NOME EMISSOR"
    sheet["C1"] = "CNPJ EMISSOR"
    sheet["D1"] = "NOME COORDENADOR LIDER"
    sheet["E1"] = "CNPJ COORDENADOR LIDER"
    sheet["F1"] = "TIPO REQUERIMENTO"
    sheet["G1"] = "TIPO DE OFERTA"
    sheet["H1"] = "STATUS"
    sheet["I1"] = "NUMERO PROCESSO"
    sheet["J1"] = "NUMERO REGISTROS"
    sheet["K1"] = "VALOR EM REAIS"
    sheet["L1"] = "POSSUI BOOK"
    sheet["M1"] = "DATA"
    sheet["N1"] = "LINK"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Azul escuro
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")  # Texto branco e negrito
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    sheet.row_dimensions[1].height = 28
    print("Exporting...")
    for offer in data.get("registros", []):
        row = [
            offer.get("idRequerimento"),
            offer.get("nomeEmissor"),
            offer.get("cnpjEmissor"),
            offer.get("nomeCoordenadorLider"),
            offer.get("cnpjCoordenadorLider"),
            offer.get("nomeTipoRequerimento"),
            offer.get("nomeValorMobiliario"),
            offer.get("statusDaOferta"),
            offer.get("numeroProcesso"),
            offer.get("numeroRegistro"),
            offer.get("valorTotalEmReais"),
            offer.get("possuiBook"),
            offer.get("data"),
            f"https://web.cvm.gov.br/sre-publico-cvm/#/oferta-publica/{offer.get("idRequerimento")}"
        ]
        sheet.append(row)
    for col in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
    wb.save("offers.xlsx")
    print("Saved offers.xlsx")